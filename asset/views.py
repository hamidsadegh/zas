import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views.generic import CreateView, DeleteView, UpdateView

from openpyxl import Workbook, load_workbook

from asset.forms import InventoryItemForm
from asset.models import InventoryItem
from dcim.models import Area, Site, Vendor


PER_PAGE_OPTIONS = (10, 25, 50, 100, 200)
QUICK_FILTER_PARAMS = (
    "qf_designation",
    "qf_inventory_number",
    "qf_serial_number",
    "qf_location",
    "qf_vendor",
    "qf_model",
    "qf_item_type",
    "qf_status",
    "qf_comment",
)


def _natural_sort_key(value):
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", value or "")
    ]


def _get_paginate_by(request, default=25):
    per_page = request.GET.get("paginate_by")
    if per_page and per_page.isdigit():
        per_page = int(per_page)
        if per_page in PER_PAGE_OPTIONS:
            return per_page
    return default


def _contains(value, query):
    return query.lower() in (value or "").lower()


def _clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _header_index_map(header_row):
    header_map = {}
    for idx, value in enumerate(header_row):
        key = _normalize_header(_clean_cell(value))
        if key:
            header_map[key] = idx
    return header_map


def _header_index(header_map, aliases):
    for alias in aliases:
        idx = header_map.get(alias)
        if idx is not None:
            return idx
    return None


def _row_cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return _clean_cell(row[idx])


def _format_validation_error(error):
    if hasattr(error, "message_dict"):
        parts = []
        for key, values in error.message_dict.items():
            joined = ", ".join(str(value) for value in values)
            parts.append(f"{key}: {joined}")
        return "; ".join(parts)
    if hasattr(error, "messages"):
        return "; ".join(str(msg) for msg in error.messages)
    return str(error)


def _resolve_site(site_name, site_id, location):
    if site_id:
        site = Site.objects.filter(id=site_id).first()
        if not site:
            raise ValueError(f"Site ID '{site_id}' does not exist.")
        return site

    resolved_site_name = site_name
    if not resolved_site_name and location:
        if " / " in location:
            resolved_site_name = location.split(" / ", 1)[0].strip()
        else:
            resolved_site_name = location.strip()

    if not resolved_site_name:
        raise ValueError("Site is required (provide Site or Site ID column).")

    matches = list(Site.objects.filter(name__iexact=resolved_site_name).order_by("name"))
    if not matches:
        raise ValueError(f"Site '{resolved_site_name}' does not exist.")
    if len(matches) > 1:
        raise ValueError(
            f"Site '{resolved_site_name}' is ambiguous. Please provide Site ID."
        )
    return matches[0]


def _resolve_area(site, area_name, area_id):
    if area_id:
        area = Area.objects.filter(id=area_id, site=site).first()
        if not area:
            raise ValueError(
                f"Area ID '{area_id}' does not exist for site '{site.name}'."
            )
        return area

    if not area_name:
        return None

    site_areas = list(Area.objects.filter(site=site))
    normalized = area_name.lower()
    matches = [
        area
        for area in site_areas
        if area.name.lower() == normalized or str(area).lower() == normalized
    ]
    if not matches:
        raise ValueError(f"Area '{area_name}' does not exist in site '{site.name}'.")
    if len(matches) > 1:
        raise ValueError(
            f"Area '{area_name}' is ambiguous in site '{site.name}'. Please use Area ID."
        )
    return matches[0]


def _resolve_vendor(vendor_name):
    if not vendor_name:
        return None
    vendor = Vendor.objects.filter(name__iexact=vendor_name).first()
    if not vendor:
        raise ValueError(f"Vendor '{vendor_name}' does not exist.")
    return vendor


def _resolve_choice(value, choices, default, field_label):
    if not value:
        return default

    normalized = value.lower()
    mapping = {}
    for internal, display in choices:
        mapping[str(internal).lower()] = internal
        mapping[str(display).lower()] = internal

    resolved = mapping.get(normalized)
    if resolved is None:
        raise ValueError(f"Invalid {field_label} '{value}'.")
    return resolved


def _storage_site_choices():
    choices = [("all", "All Sites")]
    for site in Site.objects.order_by("name"):
        choices.append((str(site.id), site.name))
    return choices


def _storage_quick_filter_values(request):
    return {
        key: request.GET.get(key, "").strip()
        for key in QUICK_FILTER_PARAMS
    }


def _build_storage_rows(search_query):
    queryset = (
        InventoryItem.objects.select_related("vendor", "site", "area")
        .all()
    )

    query = (search_query or "").lower().strip()
    rows = []

    for item in queryset:
        location = str(item.area) if item.area else item.site.name if item.site else ""
        haystack = " ".join(
            filter(
                None,
                [
                    item.designation,
                    item.inventory_number or "",
                    item.serial_number or "",
                    item.vendor.name if item.vendor else "",
                    item.model,
                    item.site.name if item.site else "",
                    str(item.area) if item.area else "",
                    item.get_item_type_display(),
                    item.get_status_display(),
                    item.comment or "",
                ],
            )
        ).lower()
        if query and query not in haystack:
            continue

        rows.append(
            {
                "item": item,
                "designation": item.designation or "",
                "inventory_number": item.inventory_number or "",
                "serial_number": item.serial_number or "",
                "location": location,
                "vendor_name": item.vendor.name if item.vendor else "",
                "model": item.model or "",
                "item_type": item.get_item_type_display(),
                "status": item.get_status_display(),
                "comment": item.comment or "",
                "site_id": str(item.site.id) if item.site else "",
                "site_name": item.site.name if item.site else "",
                "area_id": str(item.area.id) if item.area else "",
                "area_name": str(item.area) if item.area else "",
            }
        )

    return rows


def _sort_storage_rows(rows, sort_field):
    reverse = False
    field = sort_field or "designation"
    if field.startswith("-"):
        reverse = True
        field = field[1:]

    sort_keys = {
        "designation": lambda row: _natural_sort_key(row["designation"]),
        "inventory_number": lambda row: _natural_sort_key(row["inventory_number"]),
        "serial_number": lambda row: str(row["serial_number"]).lower(),
        "location": lambda row: _natural_sort_key(row["location"]),
        "vendor_name": lambda row: _natural_sort_key(row["vendor_name"]),
        "model": lambda row: _natural_sort_key(row["model"]),
        "item_type": lambda row: _natural_sort_key(row["item_type"]),
        "status": lambda row: _natural_sort_key(row["status"]),
        "comment": lambda row: str(row["comment"]).lower(),
    }

    sort_key = sort_keys.get(field, sort_keys["designation"])
    rows.sort(key=sort_key, reverse=reverse)
    return rows


def _storage_view_data(request):
    search_query = request.GET.get("search", "").strip()
    sort_field = request.GET.get("sort", "designation")
    quick_filter_values = _storage_quick_filter_values(request)

    site_choices = _storage_site_choices()
    site_filter = request.GET.get("site", "all")
    valid_site_keys = {choice[0] for choice in site_choices}
    if site_filter not in valid_site_keys:
        site_filter = "all"

    rows = _build_storage_rows(search_query)

    if site_filter != "all":
        rows = [row for row in rows if row["site_id"] == site_filter]

    if quick_filter_values["qf_designation"]:
        rows = [
            row
            for row in rows
            if _contains(row["designation"], quick_filter_values["qf_designation"])
        ]
    if quick_filter_values["qf_inventory_number"]:
        rows = [
            row
            for row in rows
            if _contains(
                row["inventory_number"], quick_filter_values["qf_inventory_number"]
            )
        ]
    if quick_filter_values["qf_serial_number"]:
        rows = [
            row
            for row in rows
            if _contains(row["serial_number"], quick_filter_values["qf_serial_number"])
        ]
    if quick_filter_values["qf_location"]:
        rows = [
            row
            for row in rows
            if _contains(row["location"], quick_filter_values["qf_location"])
        ]
    if quick_filter_values["qf_vendor"]:
        rows = [
            row
            for row in rows
            if _contains(row["vendor_name"], quick_filter_values["qf_vendor"])
        ]
    if quick_filter_values["qf_model"]:
        rows = [
            row
            for row in rows
            if _contains(row["model"], quick_filter_values["qf_model"])
        ]
    if quick_filter_values["qf_item_type"]:
        rows = [
            row
            for row in rows
            if _contains(row["item_type"], quick_filter_values["qf_item_type"])
        ]
    if quick_filter_values["qf_status"]:
        rows = [
            row
            for row in rows
            if _contains(row["status"], quick_filter_values["qf_status"])
        ]
    if quick_filter_values["qf_comment"]:
        rows = [
            row
            for row in rows
            if _contains(row["comment"], quick_filter_values["qf_comment"])
        ]

    rows = _sort_storage_rows(rows, sort_field)

    return {
        "rows": rows,
        "search_query": search_query,
        "sort_field": sort_field,
        "site_choices": site_choices,
        "site_filter": site_filter,
        "quick_filter_values": quick_filter_values,
    }


@login_required
def storage_inventory_list(request):
    view_data = _storage_view_data(request)
    paginate_by = _get_paginate_by(request, default=50)
    paginator = Paginator(view_data["rows"], paginate_by)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        **view_data,
        "rows": page_obj,
        "paginate_by": paginate_by,
        "per_page_options": PER_PAGE_OPTIONS,
    }
    return render(request, "asset/storage_inventory.html", context)


@login_required
def storage_inventory_export(request):
    view_data = _storage_view_data(request)
    rows = view_data["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Storage Inventory"
    headers = [
        "Designation",
        "Inventory Number",
        "Serial Number",
        "Site",
        "Site ID",
        "Area",
        "Area ID",
        "Vendor",
        "Model",
        "Item Type",
        "Status",
        "Comment",
    ]
    ws.append(headers)

    for row in rows:
        ws.append(
            [
                row["designation"],
                row["inventory_number"],
                row["serial_number"],
                row["site_name"],
                row["site_id"],
                row["area_name"],
                row["area_id"],
                row["vendor_name"],
                row["model"],
                row["item_type"],
                row["status"],
                row["comment"],
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="storage-inventory.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def storage_inventory_import(request):
    if request.method != "POST":
        return redirect("inventory_storage")

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        messages.error(request, "Please choose an Excel file to import.")
        return redirect("inventory_storage")

    try:
        workbook = load_workbook(excel_file, data_only=True)
    except Exception as exc:
        messages.error(request, f"Import failed: invalid Excel file ({exc}).")
        return redirect("inventory_storage")

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        messages.error(request, "Import failed: Excel sheet is empty.")
        return redirect("inventory_storage")

    header_map = _header_index_map(header_row)
    idx = {
        "designation": _header_index(header_map, ("designation",)),
        "inventory_number": _header_index(
            header_map, ("inventorynumber", "inventorynr", "inventoryno")
        ),
        "serial_number": _header_index(header_map, ("serialnumber", "serialno")),
        "site": _header_index(header_map, ("site", "sitename")),
        "site_id": _header_index(header_map, ("siteid",)),
        "area": _header_index(header_map, ("area", "location", "areaname")),
        "area_id": _header_index(header_map, ("areaid",)),
        "vendor": _header_index(header_map, ("vendor",)),
        "model": _header_index(header_map, ("model",)),
        "item_type": _header_index(header_map, ("itemtype", "type")),
        "status": _header_index(header_map, ("status",)),
        "comment": _header_index(header_map, ("comment", "notes", "description")),
    }

    missing_required = []
    if idx["designation"] is None:
        missing_required.append("Designation")
    if idx["model"] is None:
        missing_required.append("Model")
    if idx["site"] is None and idx["site_id"] is None:
        missing_required.append("Site or Site ID")

    if missing_required:
        messages.error(
            request,
            "Import failed: missing required columns: "
            + ", ".join(missing_required),
        )
        return redirect("inventory_storage")

    created_count = 0
    updated_count = 0
    skipped_count = 0
    errors = []

    seen_inventory_numbers = {}
    seen_serial_numbers = {}

    try:
        with transaction.atomic():
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not any(_clean_cell(value) for value in row):
                    continue

                designation_raw = _row_cell(row, idx["designation"])
                model = _row_cell(row, idx["model"])
                inventory_number = _row_cell(row, idx["inventory_number"]) or None
                serial_number = _row_cell(row, idx["serial_number"]) or None
                site_name = _row_cell(row, idx["site"])
                site_id = _row_cell(row, idx["site_id"])
                area_name = _row_cell(row, idx["area"])
                area_id = _row_cell(row, idx["area_id"])
                vendor_name = _row_cell(row, idx["vendor"])
                item_type_raw = _row_cell(row, idx["item_type"])
                status_raw = _row_cell(row, idx["status"])
                comment = _row_cell(row, idx["comment"])

                if not designation_raw:
                    errors.append(f"Row {row_number}: Designation is required.")
                    continue
                if not model:
                    errors.append(f"Row {row_number}: Model is required.")
                    continue

                if inventory_number:
                    previous_row = seen_inventory_numbers.get(inventory_number)
                    if previous_row is not None:
                        errors.append(
                            f"Row {row_number}: Inventory number '{inventory_number}' "
                            f"is duplicated in file (already used in row {previous_row})."
                        )
                        continue
                    seen_inventory_numbers[inventory_number] = row_number

                if serial_number:
                    previous_row = seen_serial_numbers.get(serial_number)
                    if previous_row is not None:
                        errors.append(
                            f"Row {row_number}: Serial number '{serial_number}' "
                            f"is duplicated in file (already used in row {previous_row})."
                        )
                        continue
                    seen_serial_numbers[serial_number] = row_number

                try:
                    designation = _resolve_choice(
                        designation_raw,
                        InventoryItem.Designation.choices,
                        None,
                        "designation",
                    )
                    if not designation:
                        raise ValueError("Designation is required.")
                    site = _resolve_site(site_name, site_id, area_name)
                    area = _resolve_area(site, area_name, area_id)
                    vendor = _resolve_vendor(vendor_name)
                    item_type = _resolve_choice(
                        item_type_raw,
                        InventoryItem.ItemType.choices,
                        InventoryItem.ItemType.DEVICE,
                        "item type",
                    )
                    status = _resolve_choice(
                        status_raw,
                        InventoryItem.Status.choices,
                        InventoryItem.Status.IN_STOCK,
                        "status",
                    )
                except ValueError as exc:
                    errors.append(f"Row {row_number}: {exc}")
                    continue

                existing_by_inventory = (
                    InventoryItem.objects.filter(inventory_number=inventory_number).first()
                    if inventory_number
                    else None
                )
                existing_by_serial = (
                    InventoryItem.objects.filter(serial_number=serial_number).first()
                    if serial_number
                    else None
                )

                if (
                    existing_by_inventory
                    and existing_by_serial
                    and existing_by_inventory.pk != existing_by_serial.pk
                ):
                    errors.append(
                        f"Row {row_number}: Inventory number '{inventory_number}' and "
                        f"serial number '{serial_number}' refer to different records."
                    )
                    continue

                item = existing_by_inventory or existing_by_serial or InventoryItem()
                is_create = item.pk is None

                item.designation = designation
                item.inventory_number = inventory_number
                item.serial_number = serial_number
                item.vendor = vendor
                item.model = model
                item.site = site
                item.area = area
                item.comment = comment
                item.item_type = item_type
                item.status = status

                try:
                    item.full_clean()
                    item.save()
                except ValidationError as exc:
                    errors.append(
                        f"Row {row_number}: {_format_validation_error(exc)}"
                    )
                    continue

                if is_create:
                    created_count += 1
                else:
                    updated_count += 1

            if errors:
                message = "Import aborted. " + " | ".join(errors[:12])
                if len(errors) > 12:
                    message += f" | ... and {len(errors) - 12} more errors."
                raise ValueError(message)

    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("inventory_storage")

    if created_count == 0 and updated_count == 0:
        skipped_count = 1

    success_message = (
        f"Import completed. Created: {created_count}, Updated: {updated_count}."
    )
    if skipped_count:
        success_message += " No data rows were imported."
    messages.success(request, success_message)
    return redirect("inventory_storage")


class InventoryItemAddView(LoginRequiredMixin, CreateView):
    model = InventoryItem
    form_class = InventoryItemForm
    template_name = "asset/storage_inventory_form.html"
    success_url = reverse_lazy("inventory_storage")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_edit"] = False
        return context

    def form_valid(self, form):
        messages.success(self.request, "Storage inventory item created successfully.")
        return super().form_valid(form)


class InventoryItemUpdateView(LoginRequiredMixin, UpdateView):
    model = InventoryItem
    form_class = InventoryItemForm
    template_name = "asset/storage_inventory_form.html"
    success_url = reverse_lazy("inventory_storage")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_edit"] = True
        return context

    def form_valid(self, form):
        messages.success(self.request, "Storage inventory item updated successfully.")
        return super().form_valid(form)


class InventoryItemDeleteView(LoginRequiredMixin, DeleteView):
    model = InventoryItem
    template_name = "asset/storage_inventory_confirm_delete.html"
    success_url = reverse_lazy("inventory_storage")

    def form_valid(self, form):
        messages.success(self.request, "Storage inventory item deleted.")
        return super().form_valid(form)
