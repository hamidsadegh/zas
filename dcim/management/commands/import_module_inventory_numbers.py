from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from dcim.models import DeviceModule
from services.validation_service import normalize_serial_number


class Command(BaseCommand):
    help = (
        "Import module inventory numbers from an Excel file. "
        "By default: inventory number column C, serial number column H."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "file",
            help="Path to the Excel file (.xlsx)",
        )
        parser.add_argument(
            "--sheet",
            help="Worksheet name. If omitted, the active sheet is used.",
        )
        parser.add_argument(
            "--start-row",
            type=int,
            default=2,
            help="Row number to start reading from (default: 2).",
        )
        parser.add_argument(
            "--inventory-column",
            default="C",
            help="Column letter for inventory numbers (default: C).",
        )
        parser.add_argument(
            "--serial-column",
            default="H",
            help="Column letter for serial numbers (default: H).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Preview changes without saving to the database.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"]).expanduser()
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")
        if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            raise CommandError("Only Excel OpenXML files are supported (.xlsx/.xlsm/.xltx/.xltm).")

        start_row = options["start_row"]
        if start_row < 1:
            raise CommandError("--start-row must be 1 or greater.")

        try:
            inventory_col = column_index_from_string(str(options["inventory_column"]).strip().upper())
            serial_col = column_index_from_string(str(options["serial_column"]).strip().upper())
        except ValueError as exc:
            raise CommandError(f"Invalid column letter: {exc}") from exc

        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        try:
            sheet_name = options.get("sheet")
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise CommandError(
                        f"Sheet '{sheet_name}' not found. Available: {', '.join(workbook.sheetnames)}"
                    )
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            counters = {
                "rows_scanned": 0,
                "updated": 0,
                "unchanged": 0,
                "not_found": 0,
                "skipped": 0,
            }

            self.stdout.write(
                f"Processing file: {file_path} | sheet: {sheet.title} | "
                f"inventory column: {options['inventory_column']} | serial column: {options['serial_column']}"
            )

            for row_number in range(start_row, sheet.max_row + 1):
                counters["rows_scanned"] += 1
                serial_raw = sheet.cell(row=row_number, column=serial_col).value
                inventory_raw = sheet.cell(row=row_number, column=inventory_col).value

                serial = normalize_serial_number(serial_raw)
                inventory_number = self._normalize_text(inventory_raw)

                if not serial or not inventory_number:
                    counters["skipped"] += 1
                    continue

                module = DeviceModule.objects.filter(serial_number__iexact=serial).first()
                if not module:
                    counters["not_found"] += 1
                    continue

                current_inventory = self._normalize_text(module.inventory_number)
                if current_inventory == inventory_number:
                    counters["unchanged"] += 1
                    continue

                counters["updated"] += 1
                if not options["dry_run"]:
                    module.inventory_number = inventory_number
                    module.save(update_fields=["inventory_number"])

            summary = (
                f"Completed. Rows scanned: {counters['rows_scanned']}, "
                f"updated: {counters['updated']}, unchanged: {counters['unchanged']}, "
                f"not found: {counters['not_found']}, skipped: {counters['skipped']}."
            )
            if options["dry_run"]:
                self.stdout.write(self.style.WARNING("Dry-run mode: no changes were saved."))
            self.stdout.write(self.style.SUCCESS(summary))
        finally:
            workbook.close()

    @staticmethod
    def _normalize_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()
