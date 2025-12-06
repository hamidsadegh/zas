from io import BytesIO

import pytest # type: ignore
from django.urls import reverse # type: ignore
from openpyxl import load_workbook # type: ignore
from django.contrib.auth.models import User

from dcim.models import Organization, Site, VLAN


def _login(client):
    user = User.objects.create_user(username="vlantester", password="pass")
    client.force_login(user)
    return user


def _create_vlan(site, **overrides):
    defaults = {
        "vlan_id": 10,
        "name": "Core",
        "subnet": "10.0.0.0/24",
        "gateway": "10.0.0.1",
        "usage_area": "ACI",
        "description": "",
        "site": site,
    }
    defaults.update(overrides)
    return VLAN.objects.create(**defaults)


@pytest.mark.django_db
def test_vlan_list_view_filters_by_site_and_search(client):
    _login(client)
    org = Organization.objects.create(name="Org")
    berlin = Site.objects.create(name="Berlin", organization=org)
    bonn = Site.objects.create(name="Bonn", organization=org)
    _create_vlan(vlan_id=10, name="Core VLAN Berlin", site=berlin)
    _create_vlan(vlan_id=11, name="Core VLAN Bonn", site=bonn)
    _create_vlan(vlan_id=12, name="Edge VLAN", site=berlin)

    response = client.get(reverse("vlan_list"), {"site": str(berlin.id), "q": "core"})

    assert response.status_code == 200
    vlans = list(response.context["vlans"])
    assert len(vlans) == 1
    assert vlans[0].name == "Core VLAN Berlin"
    assert response.context["site_filter"] == str(berlin.id)
    assert response.context["search_query"] == "core"


@pytest.mark.django_db
def test_vlan_list_view_respects_paginate_by(client):
    _login(client)
    org = Organization.objects.create(name="Org")
    berlin = Site.objects.create(name="Berlin", organization=org)
    for idx in range(30):
        _create_vlan(vlan_id=100 + idx, name=f"VLAN {idx}", site=berlin)

    response = client.get(reverse("vlan_list"), {"site": str(berlin.id), "paginate_by": "10"})

    assert response.status_code == 200
    assert response.context["paginator"].per_page == 10
    assert response.context["paginate_by_value"] == 10
    assert response.context["is_paginated"] is True
    assert len(response.context["vlans"]) == 10


@pytest.mark.django_db
def test_vlan_export_view_filters_by_site(client):
    _login(client)
    org = Organization.objects.create(name="Org")
    berlin = Site.objects.create(name="Berlin", organization=org)
    bonn = Site.objects.create(name="Bonn", organization=org)
    _create_vlan(vlan_id=50, name="Berlin Export VLAN", site=berlin, description="Berlin VLAN")
    _create_vlan(vlan_id=60, name="Bonn VLAN", site=bonn, description="Bonn VLAN")

    response = client.get(reverse("vlan_export"), {"site": str(berlin.id)})

    assert response.status_code == 200
    assert response["Content-Disposition"] == 'attachment; filename="vlans.xlsx"'
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("Site", "VLAN ID", "Name", "Subnet", "Gateway", "Usage Area", "Description")
    assert len(rows) == 2
    assert rows[1][0] == "Berlin"
    assert rows[1][1] == 50
    assert rows[1][2] == "Berlin Export VLAN"
