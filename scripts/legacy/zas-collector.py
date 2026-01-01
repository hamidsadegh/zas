from __init__ import *
from netmiko import ConnectHandler
from openpyxl import Workbook, load_workbook
import socket
import textfsm
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from io import StringIO
from pathlib import Path
import threading
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


@dataclass
class DeviceResult:
    device: str
    success: bool
    rest_only: bool = False
    error: Optional[str] = None
    device_type: Optional[str] = None
    hostname: Optional[str] = None
    model: Optional[str] = None
    ip_address: Optional[str] = None
    location: Optional[Tuple[str, Optional[str]]] = None
    location_message: Optional[str] = None
    inventory_rows: List[Tuple[str, str, str]] = field(default_factory=list)
    transceiver_rows: List[Tuple[str, str]] = field(default_factory=list)
    software_version: Optional[str] = None
    last_reboot: Optional[str] = None
    memory_percentage: Optional[str] = None
    status_mode: bool = False
    rebooted: bool = False


class DeviceInventory:
    """Gather device inventory/status information via Netmiko and export to Excel."""

    _REST_ONLY_HOSTS = {
        'bapic1-a324-44.dwelle.de',
        'bapic2-a587-c6.dwelle.de',
    }

    def __init__(self, *, max_workers: int = 10):
        self.username: Optional[str] = None
        self.password: Optional[str] = None
        self.host_ini: Optional[str] = None
        self.status: Optional[str] = None
        self.work_book: Optional[Workbook] = None
        self.worksheet = None
        self.excel_file_to_save: Optional[str] = None
        self.row = 1

        self.templates_dir = Path(get_project_dir('templates'))
        self.ntc_template_dir = Path(get_project_dir('ntc-template'))
        self.store_dir = Path(get_project_dir('store'))

        self.device_list_status_template = self.templates_dir / 'switch_list_status_template.xlsx'
        self.device_list_inventory_template = self.templates_dir / 'switch_list_inventory_template.xlsx'
        self.re_excel_file = self.templates_dir / 'rufe2025.xlsx'
        self.host_plus_file = self.store_dir / 'plus_host.ini'
        self.host_minus_file = self.store_dir / 'minus_host.ini'

        self.re_berlin: List[str] = []
        self.device_list_failed: List[str] = []
        self.rebooted_device_list: List[str] = []
        self.lst_plus: List[str] = []
        self.lst_minus: List[str] = []
        self.rest_list: List[str] = []
        self.device_list: List[str] = []

        self._textfsm_cache: Dict[Path, str] = {}
        self.max_workers = max_workers

    def help(self):
        print(r'''
        Generate status as well as inventory list of given active devices.

        Syntax:  /usr/bin/python zas-collector.py -n /Device/list/new_host.ini -u username [-p password] [-s] ...
        
        Example: /usr/bin/python zas-collector.py -n /Device/list/new_host.ini -u username -p
        
        args:
        -h --help           Help.
        -v --version        Shows version.        
        -n --new-list       Enter the file path containing new list of active devices.
        -u --username       Enter ssh remote user.
        -p --password       Enter ssh pasword. It will be asked for in the command line.
                            You do NOT have to enter the password in the command line.
                            With no -p will try to get the password from the env variable.
        -s --status         Generate a simple inventory list of devicees with their status.

        © 2023 Hamid Sadeghian
        ''')

    # ------------------------------------------------------------------ #
    # Helper methods
    # ------------------------------------------------------------------ #
    def _load_textfsm(self, template_name: str) -> textfsm.TextFSM:
        """Load and cache a TextFSM template."""
        template_path = self.ntc_template_dir / template_name
        try:
            template_text = self._textfsm_cache.get(template_path)
            if template_text is None:
                template_text = template_path.read_text()
                self._textfsm_cache[template_path] = template_text
            return textfsm.TextFSM(StringIO(template_text))
        except IOError as exc:
            msg = f'Error occurred while opening file: {template_path}'
            print('\n' + '\033[31m' + f'ERROR: {msg}' + '\033[0m' + '\n')
            zas_logger.logger.error(f'{msg} ({exc})')
            raise

    def _parse_textfsm(self, template_name: str, output: str) -> List[Sequence[str]]:
        template = self._load_textfsm(template_name)
        return template.ParseText(output)

    @staticmethod
    def _strip_lines(lines: Iterable[str]) -> List[str]:
        return [line.strip() for line in lines if line.strip()]

    def _read_lines(self, path: Path) -> List[str]:
        try:
            return self._strip_lines(path.read_text(encoding='utf-8').splitlines())
        except OSError as exc:
            print('\n' + '\033[31m' + f'ERROR: Occurred while opening file: {path}' + '\033[0m' + '\n')
            zas_logger.logger.error(f'Error occurred while opening file: {path} ({exc})')
            sys.exit(1)

    def _worksheet_set(self, row: int, col: int, value):
        self.worksheet.cell(row=row, column=col).value = value

    def _username_for_device(self, device: str) -> str:
        if 'leaf' in device or 'spine' in device:
            return f'apic#ISE\\{self.username}'
        return self.username

    def _connect_device(self, device: str, username: str) -> Tuple[ConnectHandler, str, str]:
        """Detect device type and return an active connection with show version output."""
        errors = []
        for device_type in ('cisco_nxos', 'cisco_ios'):
            try:
                conn = ConnectHandler(
                    device_type=device_type,
                    host=device,
                    username=username,
                    password=self.password,
                    fast_cli=True,
                )
                version_output = conn.send_command('show version', use_textfsm=False)
                detected_type = 'cisco_nxos' if 'NX-OS' in version_output else 'cisco_ios'
                if detected_type != device_type:
                    conn.disconnect()
                    conn = ConnectHandler(
                        device_type=detected_type,
                        host=device,
                        username=username,
                        password=self.password,
                        fast_cli=True,
                    )
                    version_output = conn.send_command('show version', use_textfsm=False)
                return conn, detected_type, version_output
            except BaseException as exc:
                errors.append(exc)
                zas_logger.logger.error(f'Connection attempt ({device_type}) failed for {device}: {exc}')
        raise errors[-1]

    def _fetch_location_values(
        self,
        device: str,
        connection: ConnectHandler,
        device_type: str,
        version_data: Sequence[str],
    ) -> Tuple[Optional[Tuple[str, Optional[str]]], Optional[str]]:
        if device_type == 'cisco_ios':
            command = 'show snmp location'
            template_name = 'cisco_ios_show_snmp_location.textfsm'
        else:
            command = 'show snmp'
            if len(version_data) > 3 and 'aci' in version_data[3]:
                template_name = 'cisco_aci_show_snmp_location.textfsm'
            else:
                template_name = 'cisco_nxos_show_snmp_location.textfsm'

        output = connection.send_command(command, use_textfsm=False)
        try:
            location_rows = self._parse_textfsm(template_name, output)
        except Exception:
            location_rows = []

        location: Optional[Tuple[str, Optional[str]]] = None
        message: Optional[str] = None

        if location_rows:
            try:
                location = (location_rows[0][0], location_rows[0][1])
            except Exception:
                try:
                    location = (location_rows[0], None)
                except Exception:
                    pass

        if len(version_data) > 3 and 'aci' in version_data[3]:
            location_parts = device.replace('.dwelle.de', '').split('-')
            if 'tsu' in device:
                location = ('TSU', 'Rack125')
            elif len(location_parts) >= 3:
                location = (location_parts[1].upper(), f'Rack{location_parts[2]}')
                message = None

        if location is None:
            message = "Bitte 'SNMP location' setzen!"

        return location, message

    def _collect_inventory_rows(
        self, device_type: str, connection: ConnectHandler
    ) -> Tuple[List[Tuple[str, str, str]], List[Tuple[str, str]]]:
        inventory_rows: List[Tuple[str, str, str]] = []
        command = 'show inventory'
        inventory_output = connection.send_command(command, use_textfsm=False)
        template_name = f'{device_type}_show_inventory.textfsm'
        try:
            parsed_inventory = self._parse_textfsm(template_name, inventory_output)
        except Exception:
            parsed_inventory = []

        for entry in parsed_inventory:
            module_name = entry[0] if len(entry) > 0 else ''
            module_desc = entry[1] if len(entry) > 1 else ''
            module_serial = entry[4] if len(entry) > 4 else ''
            inventory_rows.append((module_name, module_desc, module_serial))

        transceiver_rows: List[Tuple[str, str]] = []
        if device_type == 'cisco_nxos':
            trx_output = connection.send_command('show interface transceiver', use_textfsm=False)
            template_name = 'cisco_nxos_show_interface_transiver_detail.textfsm'
            try:
                parsed_trx = self._parse_textfsm(template_name, trx_output)
            except Exception:
                parsed_trx = []
            for entry in parsed_trx:
                module_type = entry[0] if entry else ''
                module_serial = entry[1] if len(entry) > 1 else ''
                transceiver_rows.append((module_type, module_serial))

        return inventory_rows, transceiver_rows

    def _calculate_memory_percentage(self, connection: ConnectHandler, device_type: str) -> Optional[str]:
        if device_type == 'cisco_ios':
            command = 'show memory statistics'
            template_name = 'cisco_ios_show_memory_statistics.textfsm'
        else:
            command = 'show system resources'
            template_name = 'cisco_nxos_show_system_resources.textfsm'

        output = connection.send_command(command, use_textfsm=False)
        try:
            memory_rows = self._parse_textfsm(template_name, output)
        except Exception:
            return None
        if not memory_rows:
            return None
        total = int(memory_rows[0][0])
        used = int(memory_rows[0][1])
        percentage = round((100 * used) / total, 2)
        return f'{percentage}%'

    def _collect_status_details(
        self, device_type: str, connection: ConnectHandler, version_data: Sequence[str]
    ) -> Tuple[Optional[str], Optional[str], Optional[str], bool]:
        if device_type == 'cisco_nxos':
            software_version = version_data[2]
            last_reboot = version_data[0]
            rebooted = bool(version_data[7] and int(version_data[7]) < 8)
        else:
            software_version = version_data[0]
            last_reboot = version_data[9]
            rebooted = bool(
                (not version_data[10] and not version_data[11] and not version_data[12])
                or (
                    not version_data[10]
                    and not version_data[11]
                    and version_data[12]
                    and int(version_data[12]) < 8
                )
            )

        memory_percentage = self._calculate_memory_percentage(connection, device_type)
        return software_version, last_reboot, memory_percentage, rebooted

    def _collect_device_data(self, device: str) -> DeviceResult:
        thread_name = threading.current_thread().name
        zas_logger.logger.info(f'Thread {thread_name} started collection for {device}')

        if device in self._REST_ONLY_HOSTS:
            return DeviceResult(device=device, success=False, rest_only=True)

        username = self._username_for_device(device)
        connection: Optional[ConnectHandler] = None
        try:
            connection, device_type, version_output = self._connect_device(device, username)
            version_rows = self._parse_textfsm(f'{device_type}_show_version.textfsm', version_output)
            if not version_rows:
                raise RuntimeError('Unable to parse show version output.')
            version_data = version_rows[0]

            print('\033[32m' + f'Connected to device {device}' + '\033[0m')
            zas_logger.logger.info(f'Connected to device {device}')

            ip_address = self.nslookup(device)
            if device_type == 'cisco_ios':
                hostname = version_data[2]
                model = version_data[5][0]
            else:
                hostname = version_data[5]
                model = version_data[4]

            location, location_message = self._fetch_location_values(device, connection, device_type, version_data)

            result = DeviceResult(
                device=device,
                success=True,
                device_type=device_type,
                hostname=hostname,
                model=model,
                ip_address=ip_address,
                location=location,
                location_message=location_message,
                status_mode=self.status is not None,
            )

            if self.status is None:
                inventory_rows, transceiver_rows = self._collect_inventory_rows(device_type, connection)
                result.inventory_rows = inventory_rows
                result.transceiver_rows = transceiver_rows
            else:
                (
                    result.software_version,
                    result.last_reboot,
                    result.memory_percentage,
                    result.rebooted,
                ) = self._collect_status_details(device_type, connection, version_data)

            return result
        except BaseException as exc:
            zas_logger.logger.error(f'Error collecting data for {device}: {exc}')
            return DeviceResult(device=device, success=False, error=str(exc))
        finally:
            if connection:
                try:
                    connection.disconnect()
                except Exception:
                    pass

    def _apply_device_result(self, result: DeviceResult):
        if result.rest_only:
            self.rest_list.append(result.device)
            return

        if not result.success:
            self.device_list_failed.append(result.device)
            self.rest_list.append(result.device)
            if result.error:
                zas_logger.logger.error(f'Device {result.device} failed: {result.error}')
                print('\n' + '\033[91m' + f'ERROR: {result.device} {result.error}' + '\033[0m' + '\n')
            return

        self.row += 1
        self._worksheet_set(self.row, 1, result.hostname)
        self._worksheet_set(self.row, 2, result.ip_address)
        self._worksheet_set(self.row, 3, result.model)

        if result.location:
            self._worksheet_set(self.row, 4, result.location[0])
            if result.location[1] is not None:
                self._worksheet_set(self.row, 5, result.location[1])
        elif result.location_message:
            self._worksheet_set(self.row, 4, result.location_message)

        if result.status_mode:
            self._worksheet_set(self.row, 7, 'OK')
            self._worksheet_set(self.row, 8, 'OK')
            if result.software_version:
                self._worksheet_set(self.row, 6, result.software_version)
            if result.last_reboot:
                self._worksheet_set(self.row, 11, result.last_reboot)
            if result.memory_percentage:
                self._worksheet_set(self.row, 9, result.memory_percentage)
            if result.rebooted:
                self.rebooted_device_list.append(result.device)
        else:
            for module_name, module_desc, module_serial in result.inventory_rows:
                self.row += 1
                self._worksheet_set(self.row, 6, module_name)
                self._worksheet_set(self.row, 7, module_desc)
                self._worksheet_set(self.row, 8, module_serial)

            for module_type, module_serial in result.transceiver_rows:
                self.row += 1
                self._worksheet_set(self.row, 7, module_type)
                self._worksheet_set(self.row, 8, module_serial)

    # ------------------------------------------------------------------ #
    # Core workflows
    # ------------------------------------------------------------------ #
    def connect_to_device_gather_information(self):
        zas_logger.logger.info(f'{self.connect_to_device_gather_information.__name__} function called')
        if not self.device_list:
            return

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            results = list(executor.map(self._collect_device_data, self.device_list))

        for result in results:
            self._apply_device_result(result)

    def rest_list_fill(self):
        zas_logger.logger.info('rest_list_fill() method is called')
        for device in self.rest_list:
            self.row += 1
            self._worksheet_set(self.row, 1, device)
            self._worksheet_set(self.row, 2, self.nslookup(device))
            for col in range(3, 12):
                self._worksheet_set(self.row, col, 'n.a.')
            if self.status == 'set':
                self._worksheet_set(self.row, 7, 'OK')
                self._worksheet_set(self.row, 8, 'n.a.')

    def on_call(self):
        zas_logger.logger.info(f'{self.on_call.__name__} function called.')
        try:
            re_work_book = load_workbook(self.re_excel_file)
            re_work_book.active = 1
            re_worksheet = re_work_book['RE']
            for column in range(5, 15):
                cell_value = re_worksheet.cell(row=int(week) + 2, column=column).value
                if cell_value is not None:
                    self.re_berlin.append(re_worksheet.cell(row=1, column=column).value)
        except IOError:
            print('\n' + '\033[91m' + f'ERROR: Occurred while opening file: {self.re_excel_file}' + '\033[0m' + '\n')
            zas_logger.logger.error(f'Error occurred while opening file: {self.re_excel_file}')

    def report(self):
        zas_logger.logger.info(f'{self.report.__name__} function called')
        report_path = self.store_dir / 'report'
        with report_path.open('w', encoding='utf-8') as f:
            f.write(f'Subject: Netzwerkstatus Berlin und Bonn {today_dot} Woche {week}')
            f.write('\nHallo zusammen,')
            f.write('\n\nBerliner und Bonner Management-Netze sind gescannt.')
            f.write('\n\nDie aktuelle Switchlisten für Berlin und Bonn sind unter folgendem Link verfügbar. ')
            f.write('https://tib.dw.com/pages/viewpage.action?pageId=254315471')
            f.write('\n\n' + f'Ruferreichbarkeit Berlin/Bonn: {self.re_berlin}')
            f.write('\n\nBitte überprüfen Sie den Bericht auf Vollständigkeit und Richtigkeit.')
            f.write('\n\nEinen guten Start in die Woche!')
        zas_logger.logger.info(f'Report file generated {report_path}')

    # ------------------------------------------------------------------ #
    # Networking helpers
    # ------------------------------------------------------------------ #
    def nslookup(self, hostname: str) -> str:
        try:
            return socket.gethostbyname(hostname)
        except socket.gaierror as exc:
            zas_logger.logger.error(f'No DNS Record: {hostname} ({exc})')
            return 'No DNS Record'

    def reverse_lookup(self, ip_address: str) -> str:
        try:
            host_name = socket.gethostbyaddr(ip_address)
            return host_name[0]
        except socket.gaierror as exc:
            zas_logger.logger.error(f'No DNS Record: {ip_address} ({exc})')
            return 'No DNS Record'

    # ------------------------------------------------------------------ #
    # CLI / main
    # ------------------------------------------------------------------ #
    def _parse_arguments(self, argv: Sequence[str]):
        try:
            opts, args = getopt.getopt(
                argv[1:],
                'u:p:i:hvs',
                ['help', 'version', 'status', 'user=', 'password', 'host-ini='],
            )
        except getopt.GetoptError as exc:
            self.help()
            print('\n' + '\033[31m' + f'ERROR: Parameter {exc}' + '\033[0m' + '\n')
            zas_logger.logger.error(f'Parameter {exc}')
            sys.exit(2)

        for opt, arg in opts:
            if opt in ('-h', '--help'):
                self.help()
                sys.exit()
            elif opt in ('-v', '--version'):
                print('Version ', version)
                sys.exit()
            elif opt in ('-u', '--user'):
                self.username = arg
            elif opt in ('-p', '--password'):
                print('User:', self.username)
                self.password = getpass.getpass()
            elif opt in ('-i', '--host-ini'):
                self.host_ini = arg
            elif opt in ('-s', '--status'):
                self.status = 'set'
            else:
                self.help()
                print('\n' + '\033[31m' + f'ERROR: Parameter {opt}' + '\033[0m' + '\n')
                zas_logger.logger.error(f'Parameter {opt}')
                sys.exit()

    def _ensure_cli_requirements(self):
        if not self.username:
            self.help()
            print('\n' + '\033[31m' + 'ERROR: Mandatory parameter is missing. (-u)' + '\033[0m' + '\n')
            zas_logger.logger.error('Mandatory parameter is missing. (-u)')
            sys.exit()

        if not self.password:
            self.password = keyring.get_password('zas', self.username)

        if not self.host_ini:
            self.help()
            print('\033[31m' + 'ERROR: Mandatory parameter is missing.' + '\033[0m' + '\n')
            zas_logger.logger.error('Mandatory parameter is missing.')
            sys.exit()

    def _load_workbook_template(self):
        self.work_book = Workbook()
        if self.status == 'set':
            try:
                self.work_book = load_workbook(self.device_list_status_template)
                self.excel_file_to_save = str(self.store_dir / 'Berlin_Switche_Status.xlsx')
            except IOError:
                print('\n' + '\033[31m' + f'ERROR: Occurred while opening file: {self.device_list_status_template}' + '\033[0m' + '\n')
                zas_logger.logger.error(f'Error occurred while opening file: {self.device_list_status_template}')
        else:
            try:
                self.work_book = load_workbook(self.device_list_inventory_template)
                self.excel_file_to_save = str(self.store_dir / 'Berlin_Switche_Inventarliste.xlsx')
            except IOError:
                print('\n' + '\033[31m' + f'ERROR: Occurred while opening file: {self.device_list_inventory_template}' + '\033[0m' + '\n')
                zas_logger.logger.error(f'Error occurred while opening file: {self.device_list_inventory_template}')

        self.work_book.active = 1
        self.worksheet = self.work_book['switche']
        self.row = 1

    def _load_device_lists(self):
        host_path = Path(self.host_ini)
        self.device_list = self._read_lines(host_path)
        self.lst_plus = self._read_lines(self.host_plus_file)
        self.lst_minus = self._read_lines(self.host_minus_file)

    def main(self, argv: Sequence[str]):
        self._parse_arguments(argv)
        self._ensure_cli_requirements()
        self._load_device_lists()
        self._load_workbook_template()

        self.connect_to_device_gather_information()
        self.rest_list_fill()
        self.on_call()
        self.report()

        self.work_book.save(self.excel_file_to_save)
        zas_logger.logger.info(f'An inventory of devicees inclusive their status stored in {self.excel_file_to_save}.')

        print('\n' + '\033[32m' + 'Done!' + '\033[0m')
        zas_logger.logger.info('Done!')


if __name__ == '__main__':
    zas_logger = LoggerConfig(os.path.basename(__file__).split(".")[0])
    zas_logger.logger.info(f'{__file__} started.')
    inventory = DeviceInventory()
    inventory.main(sys.argv)
    zas_logger.logger.info(f'{__file__} finished.')
