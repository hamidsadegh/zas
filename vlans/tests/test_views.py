from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from vlans.models import VLAN


def _create_vlan(**overrides):
    defaults = {
        "site": "Berlin",
        "vlan_id": 10,
        "name": "Core",
        "subnet": "10.0.0.0/24",
        "gateway": "10.0.0.1",
        "usage_area": "ACI",
        "description": "",
    }
    defaults.update(overrides)
    return VLAN.objects.create(**defaults)


@pytest.mark.django_db
def test_vlan_list_view_filters_by_site_and_search(client):
    _create_vlan(vlan_id=10, name="Core VLAN Berlin", site="Berlin")
    _create_vlan(vlan_id=11, name="Core VLAN Bonn", site="Bonn")
    _create_vlan(vlan_id=12, name="Edge VLAN", site="Berlin")

    response = client.get(reverse("vlan_list"), {"site": "Berlin", "q": "core"})

    assert response.status_code == 200
    vlans = list(response.context["vlans"])
    assert len(vlans) == 1
    assert vlans[0].name == "Core VLAN Berlin"
    assert response.context["site_filter"] == "Berlin"
    assert response.context["search_query"] == "core"


@pytest.mark.django_db
def test_vlan_list_view_respects_paginate_by(client):
    for idx in range(30):
        _create_vlan(vlan_id=100 + idx, name=f"VLAN {idx}")

    response = client.get(reverse("vlan_list"), {"site": "Berlin", "paginate_by": "10"})

    assert response.status_code == 200
    assert response.context["paginator"].per_page == 10
    assert response.context["paginate_by_value"] == 10
    assert response.context["is_paginated"] is True
    assert len(response.context["vlans"]) == 10


@pytest.mark.django_db
def test_vlan_export_view_filters_by_site(client):
    _create_vlan(vlan_id=50, name="Berlin Export VLAN", site="Berlin", description="Berlin VLAN")
    _create_vlan(vlan_id=60, name="Bonn VLAN", site="Bonn", description="Bonn VLAN")

    response = client.get(reverse("vlan_export"), {"site": "Berlin"})

    assert response.status_code == 200
    assert response["Content-Disposition"] == 'attachment; filename="vlans.xlsx"'
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("Site", "VLAN ID", "Name", "Subnet", "Gateway", "Usage Area", "Description")
    assert len(rows) == 2
    assert rows[1][0] == "Berlin"
    assert rows[1][1] == 50
    assert rows[1][2] == "Berlin Export VLAN"
